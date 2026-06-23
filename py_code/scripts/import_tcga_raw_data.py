#!/usr/bin/env python3
import argparse
import csv
import gzip
import io
import json
import re
import shlex
import subprocess
import sys
import tarfile
from pathlib import Path


DB_ARGS = ["docker", "exec", "-i", "bio-postgres", "psql", "-U", "bio", "-d", "bio", "-v", "ON_ERROR_STOP=1"]
SCHEMA = "bio_tcga"
MAX_WIDE_COLUMNS = 1600
MATRIX_COPY_DELIMITER = "\x1f"


DIRECT_TABLES = {
    "PanCan-General_Open_GDC-Manifest_2.txt": "gdc_manifest",
    "PanCanAtlas_miRNA_sample_information_list.txt": "mirna_sample_information",
    "TCGA-RPPA-pancan-clean.txt": "rppa_pancan_clean",
    "TCGA_mastercalls.abs_segtabs.fixed.txt": "abs_segtabs",
    "TCGA_mastercalls.abs_tables_JSedit.fixed.txt": "abs_tables",
    "broad.mit.edu_PANCAN_Genome_Wide_SNP_6_whitelisted.seg": "snp6_whitelisted_seg",
    "clinical_PANCAN_patient_with_followup.tsv": "clinical_patient_followup",
    "mc3.v0.2.8.PUBLIC.maf": "mc3_public_maf",
    "merged_sample_quality_annotations.tsv": "sample_quality_annotations",
}


MATRIX_TABLES = {
    "EBPlusPlusAdjustPANCAN_IlluminaHiSeq_RNASeqV2.geneExp.tsv": ("matrix_rnaseq_gene_expression", 1),
    "jhu-usc.edu_PANCAN_HumanMethylation450.betaValue_whitelisted.tsv": ("matrix_methylation450_beta", 1),
    "jhu-usc.edu_PANCAN_merged_HumanMethylation27_HumanMethylation450.betaValue_whitelisted.tsv": ("matrix_methylation27_450_beta", 1),
    "pancanMiRs_EBadjOnProtocolPlatformWithoutRepsWithUnCorrectMiRs_08_04_16.csv": ("matrix_mirna_ebadj", 2),
    "merge_merged_reals.tar.gz": ("matrix_merge_merged_reals", 1),
}


def qident(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def sanitize_column(name: str, used: set[str], fallback: str) -> str:
    cleaned = name.strip().strip('"')
    if not cleaned:
        cleaned = fallback
    cleaned = re.sub(r"[^0-9A-Za-z_]+", "_", cleaned).strip("_").lower()
    if not cleaned:
        cleaned = fallback
    if cleaned[0].isdigit():
        cleaned = "c_" + cleaned
    cleaned = cleaned[:55]
    base = cleaned
    i = 2
    while cleaned in used:
        suffix = f"_{i}"
        cleaned = base[: 63 - len(suffix)] + suffix
        i += 1
    used.add(cleaned)
    return cleaned


def run_sql(sql: str) -> None:
    subprocess.run(DB_ARGS, input=sql.encode(), check=True)


def copy_bytes(sql: str, payload) -> None:
    proc = subprocess.Popen(DB_ARGS + ["-c", sql], stdin=subprocess.PIPE)
    try:
        for chunk in payload:
            proc.stdin.write(chunk)
    finally:
        if proc.stdin:
            proc.stdin.close()
    if proc.wait() != 0:
        raise subprocess.CalledProcessError(proc.returncode, DB_ARGS)


def copy_tsv_matrix_with_client_container(raw_dir: Path, path: Path, table: str, copy_cols: str, tar_member: str | None) -> None:
    copy_sql = (
        f"COPY {qident(SCHEMA)}.{qident(table)} ({copy_cols}) "
        f"FROM STDIN WITH (FORMAT text, DELIMITER E'\\x1f', NULL '')"
    )
    awk = r"""awk 'BEGIN{FS="\t"; OFS=sprintf("%c",31)} NR>1{feature=$1; sub(/^[^\t]*\t/,""); print feature,$0}'"""
    mounted_file = "/raw/" + path.name
    if tar_member:
        program = f"tar -xOzf {shlex.quote(mounted_file)} {shlex.quote(tar_member)} | {awk}"
    else:
        program = f"{awk} {shlex.quote(mounted_file)}"
    shell_script = f"set -e; {program} | psql -h 127.0.0.1 -U bio -d bio -v ON_ERROR_STOP=1 -c {shlex.quote(copy_sql)}"
    subprocess.run(
        [
            "docker",
            "run",
            "--rm",
            "--network",
            "container:bio-postgres",
            "-v",
            f"{raw_dir.resolve()}:/raw:ro",
            "-e",
            "PGPASSWORD=bioanalysis",
            "bio-postgres:18",
            "sh",
            "-c",
            shell_script,
        ],
        check=True,
    )


def file_bytes(path: Path):
    with path.open("rb") as f:
        while True:
            chunk = f.read(1024 * 1024)
            if not chunk:
                break
            yield chunk


def utf8_clean_file_bytes(path: Path):
    with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
        while True:
            chunk = f.read(1024 * 1024)
            if not chunk:
                break
            yield chunk.encode("utf-8")


def text_opener(path: Path):
    if path.suffix == ".gz" and not path.name.endswith(".tar.gz"):
        return gzip.open(path, "rt", encoding="utf-8", errors="replace", newline="")
    return open(path, "rt", encoding="utf-8", errors="replace", newline="")


def detect_delimiter(header: str) -> str:
    return "\t" if header.count("\t") >= header.count(",") else ","


def read_header(path: Path, tar_member: str | None = None) -> tuple[list[str], str]:
    if tar_member:
        with tarfile.open(path, "r:gz") as tar:
            f = tar.extractfile(tar_member)
            if not f:
                raise RuntimeError(f"Cannot read {tar_member} from {path}")
            header = f.readline().decode("utf-8", errors="replace").rstrip("\r\n")
    else:
        with text_opener(path) as f:
            header = f.readline().rstrip("\r\n")
    delimiter = detect_delimiter(header)
    return next(csv.reader([header], delimiter=delimiter)), delimiter


def create_meta_table() -> None:
    run_sql(f"""
CREATE SCHEMA IF NOT EXISTS {qident(SCHEMA)};
CREATE TABLE IF NOT EXISTS {qident(SCHEMA)}.import_manifest (
  table_name text PRIMARY KEY,
  source_file text NOT NULL,
  load_mode text NOT NULL,
  loaded_at timestamptz NOT NULL DEFAULT now(),
  columns_json jsonb,
  note text
);
""")


def record_manifest(table: str, source: str, mode: str, columns: list[str], note: str = "") -> None:
    payload = json.dumps(columns, ensure_ascii=False).replace("'", "''")
    run_sql(f"""
INSERT INTO {qident(SCHEMA)}.import_manifest(table_name, source_file, load_mode, columns_json, note)
VALUES ('{table}', '{source.replace("'", "''")}', '{mode}', '{payload}'::jsonb, '{note.replace("'", "''")}')
ON CONFLICT (table_name) DO UPDATE
SET source_file = EXCLUDED.source_file,
    load_mode = EXCLUDED.load_mode,
    loaded_at = now(),
    columns_json = EXCLUDED.columns_json,
    note = EXCLUDED.note;
""")


def import_direct(path: Path, table: str) -> None:
    header, delimiter = read_header(path)
    used: set[str] = set()
    columns = [sanitize_column(c, used, f"column_{i}") for i, c in enumerate(header, 1)]
    col_defs = ",\n  ".join(f"{qident(c)} text" for c in columns)
    col_list = ", ".join(qident(c) for c in columns)
    delim = "\\t" if delimiter == "\t" else ","
    run_sql(f"""
DROP TABLE IF EXISTS {qident(SCHEMA)}.{qident(table)};
CREATE TABLE {qident(SCHEMA)}.{qident(table)} (
  {col_defs}
);
""")
    sql = (
        f"COPY {qident(SCHEMA)}.{qident(table)} ({col_list}) "
        f"FROM STDIN WITH (FORMAT csv, HEADER true, DELIMITER E'{delim}', QUOTE '\"', ESCAPE '\"', NULL '')"
    )
    print(f"loading direct table {SCHEMA}.{table} from {path.name}", flush=True)
    copy_bytes(sql, utf8_clean_file_bytes(path))
    record_manifest(table, path.name, "direct_text_columns", header)


def matrix_rows(path: Path, id_cols: int, tar_member: str | None = None):
    if tar_member:
        with tarfile.open(path, "r:gz") as tar:
            f = tar.extractfile(tar_member)
            if not f:
                raise RuntimeError(f"Cannot read {tar_member} from {path}")
            wrapper = io.TextIOWrapper(f, encoding="utf-8", errors="replace", newline="")
            yield from _matrix_rows_from_text(wrapper, id_cols)
    else:
        with text_opener(path) as f:
            yield from _matrix_rows_from_text(f, id_cols)


def _matrix_rows_from_text(f, id_cols: int):
    header_line = f.readline().rstrip("\r\n")
    delimiter = detect_delimiter(header_line)
    if delimiter == "\t" and id_cols == 1:
        sep = MATRIX_COPY_DELIMITER
        for line in f:
            line = line.rstrip("\r\n")
            if not line:
                continue
            feature_id, _, values_text = line.partition("\t")
            yield f"{feature_id}{sep}{values_text}\n".encode("utf-8")
        return
    reader = csv.reader(f, delimiter=delimiter)
    writer_buf = io.StringIO()
    writer = csv.writer(writer_buf, delimiter=MATRIX_COPY_DELIMITER, lineterminator="\n")
    for row in reader:
        if not row:
            continue
        row += [""] * max(0, id_cols - len(row))
        ids = row[:id_cols]
        values = row[id_cols:]
        writer_buf.seek(0)
        writer_buf.truncate(0)
        writer.writerow(ids + ["\t".join(values)])
        yield writer_buf.getvalue().encode("utf-8")


def import_matrix(path: Path, table: str, id_cols: int, tar_member: str | None = None) -> None:
    header, delimiter = read_header(path, tar_member)
    ids = header[:id_cols]
    samples = header[id_cols:]
    id_defs = ["feature_id text"] if id_cols == 1 else ["feature_id text", "feature_type text"]
    id_copy_cols = ["feature_id"] if id_cols == 1 else ["feature_id", "feature_type"]
    run_sql(f"""
DROP TABLE IF EXISTS {qident(SCHEMA)}.{qident(table)};
DROP TABLE IF EXISTS {qident(SCHEMA)}.{qident(table + '_samples')};
CREATE UNLOGGED TABLE {qident(SCHEMA)}.{qident(table + '_samples')} (
  sample_index integer PRIMARY KEY,
  sample_id text NOT NULL
);
CREATE UNLOGGED TABLE {qident(SCHEMA)}.{qident(table)} (
  {", ".join(id_defs)},
  values_text text
);
""")
    sample_csv = io.StringIO()
    writer = csv.writer(sample_csv, lineterminator="\n")
    for idx, sample in enumerate(samples, 1):
        writer.writerow([idx, sample])
    copy_bytes(
        f"COPY {qident(SCHEMA)}.{qident(table + '_samples')} (sample_index, sample_id) FROM STDIN WITH (FORMAT csv)",
        [sample_csv.getvalue().encode("utf-8")],
    )
    copy_cols = ", ".join(qident(c) for c in id_copy_cols + ["values_text"])
    print(f"loading matrix table {SCHEMA}.{table} from {path.name} ({len(samples)} samples)", flush=True)
    if delimiter == "\t" and id_cols == 1:
        copy_tsv_matrix_with_client_container(path.parent, path, table, copy_cols, tar_member)
    else:
        copy_bytes(
            f"COPY {qident(SCHEMA)}.{qident(table)} ({copy_cols}) FROM STDIN WITH (FORMAT csv, DELIMITER E'\\x1f', NULL '')",
            matrix_rows(path, id_cols, tar_member),
        )
    record_manifest(
        table,
        path.name if not tar_member else f"{path.name}:{tar_member}",
        "matrix_raw_text_vector",
        ids + ["values_text"],
        f"samples in {table}_samples; split values_text on tab to align with sample_index",
    )


def import_xlsx(path: Path) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX import") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet in wb.worksheets:
        base = sanitize_column(sheet.title, set(), "sheet")
        table = f"tcga_cdr_{base}"[:63]
        rows = sheet.iter_rows(values_only=True)
        try:
            header_values = next(rows)
        except StopIteration:
            continue
        used: set[str] = set()
        header = ["" if v is None else str(v) for v in header_values]
        columns = [sanitize_column(c, used, f"column_{i}") for i, c in enumerate(header, 1)]
        col_defs = ",\n  ".join(f"{qident(c)} text" for c in columns)
        col_list = ", ".join(qident(c) for c in columns)
        run_sql(f"""
DROP TABLE IF EXISTS {qident(SCHEMA)}.{qident(table)};
CREATE TABLE {qident(SCHEMA)}.{qident(table)} (
  {col_defs}
);
""")
        def payload():
            buf = io.StringIO()
            writer = csv.writer(buf, lineterminator="\n")
            for row in rows:
                values = ["" if v is None else str(v) for v in row]
                values += [""] * (len(columns) - len(values))
                buf.seek(0)
                buf.truncate(0)
                writer.writerow(values[: len(columns)])
                yield buf.getvalue().encode("utf-8")
        print(f"loading xlsx sheet {SCHEMA}.{table} from {path.name}:{sheet.title}", flush=True)
        copy_bytes(
            f"COPY {qident(SCHEMA)}.{qident(table)} ({col_list}) FROM STDIN WITH (FORMAT csv, NULL '')",
            payload(),
        )
        record_manifest(table, f"{path.name}:{sheet.title}", "xlsx_text_columns", header)


def summarize() -> None:
    run_sql(f"""
SELECT schemaname || '.' || relname AS table_name, n_live_tup::bigint AS estimated_rows
FROM pg_stat_user_tables
WHERE schemaname = '{SCHEMA}'
ORDER BY relname;
""")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--raw-dir", default="data/raw_data")
    parser.add_argument("--only-matrices", action="store_true")
    parser.add_argument("--skip-large-matrices", action="store_true")
    parser.add_argument("--start-at-matrix", help="Start matrix imports at this source filename or table name")
    args = parser.parse_args()

    raw_dir = Path(args.raw_dir)
    if not raw_dir.exists():
        print(f"Raw data directory not found: {raw_dir}", file=sys.stderr)
        return 1

    create_meta_table()
    if not args.only_matrices:
        for filename, table in DIRECT_TABLES.items():
            path = raw_dir / filename
            if path.exists():
                import_direct(path, table)
        xlsx_path = raw_dir / "TCGA-CDR-SupplementalTableS1.xlsx"
        if xlsx_path.exists():
            import_xlsx(xlsx_path)
    for filename, (table, id_cols) in MATRIX_TABLES.items():
        if args.start_at_matrix and args.start_at_matrix not in {filename, table}:
            continue
        args.start_at_matrix = None
        if args.skip_large_matrices and filename not in {"pancanMiRs_EBadjOnProtocolPlatformWithoutRepsWithUnCorrectMiRs_08_04_16.csv"}:
            continue
        path = raw_dir / filename
        if not path.exists():
            continue
        tar_member = "merge_merged_reals.txt" if filename.endswith(".tar.gz") else None
        import_matrix(path, table, id_cols, tar_member)
    summarize()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
